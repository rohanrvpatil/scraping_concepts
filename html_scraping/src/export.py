# Tabular Data: CSV, Excel (XLS/XLSX), Google Sheets
# Hierarchical/Nested Data: JSON, XML, YAML
# Database: SQL
# Big Data/Analytics: Parquet
# Reports: PDF
# Web Data: HTML

import csv

from openpyxl import Workbook, load_workbook

import json
from dataclasses import asdict


from src import (
    Product
)

def export_csv(product: Product, file_path: str = './data_exports/data.csv'):
    file_exists = False
    
    try:
        with open(file_path, 'r'):
            file_exists = True
    except FileNotFoundError:
        pass
    
    with open(file_path, mode='a', newline='') as file:
        writer = csv.writer(file)
        if not file_exists:
            writer.writerow(['link', 'name', 'product_id', 'price', 'rating'])
        writer.writerow([product.link, product.name, product.product_id, product.price, product.rating])
        

def export_xlsx(product: Product, file_path: str = './data_exports/data.xlsx'):
    file_exists = False

    try:
        load_workbook(file_path)
        file_exists = True
    except FileNotFoundError:
        pass

    if file_exists:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['link', 'name', 'product_id', 'price', 'rating'])

    sheet.append([product.link, product.name, product.product_id, product.price, product.rating])

    workbook.save(file_path)
    
def export_json(product: Product, file_path: str = './data_exports/data.json'):
    try:
        with open(file_path, 'r') as file:
            data = json.load(file)
    except FileNotFoundError:
        data = []
    
    data.append(asdict(product))
    
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)